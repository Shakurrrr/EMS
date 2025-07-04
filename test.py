import random
from pathlib import Path
import shutil
import openpyxl as xl

wb = xl.load_workbook("student_list.xlsx")
sheet = wb["Sheet1"]

# Print headers
for col in range(1, 6):
    print(sheet.cell(row=1, column=col).value)

# Loop through rows
for datas in sheet.iter_rows(min_row=2, values_only=True):
    if all(cell is None for cell in datas):
        continue  # skip empty rows

    print(f"Row data: {datas}")
    
    input_name = input("Enter the name of the student: ").strip()
    
    if input_name.lower() == "exit":
        break

    if input_name.lower() == str(datas[0]).strip().lower():
        print(f"Found at index: {datas.index(datas[0])}")
    else:
        print("Student not found.")

## test for pathlib
#path = Path()
#for file in path.glob("*.*"):
#    print(file)


## test for dice roll 
#class Dice:
#    def roll(self):
#        roll1 = random.randint(1, 6)
#        roll2 = random.randint(1, 6)
#        return roll1, roll2

#dice = Dice()
#print(dice.roll())